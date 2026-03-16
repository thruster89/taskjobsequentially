"""
SAS → DuckDB 전환 파이프라인 (범용 실행기)

JOB 파일을 지정하여 실행합니다. 각 JOB 파일은 독립적으로 정의됩니다.

실행 예시:
  python sas_to_duckdb.py --ym 202601 --job jobs/job1.py
  python sas_to_duckdb.py --ym 202601 --job jobs/job2.py --skip-load
  python sas_to_duckdb.py --ym 202601 --job jobs/job1.py jobs/job2.py jobs/job3.py
  python sas_to_duckdb.py --ym 202601 --job jobs/job1.py --stage load
  python sas_to_duckdb.py --ym 202601 --job jobs/job1.py --stage logic validate
"""

import re
import sys
import logging
import time
import argparse
import unicodedata
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
import importlib.util
from pathlib import Path
from datetime import datetime

# ══════════════════════════════════════════════
# 로거
# ══════════════════════════════════════════════
def setup_logger() -> logging.Logger:
    logger = logging.getLogger("pipeline")
    if logger.handlers:
        return logger
    logger.setLevel(logging.DEBUG)

    class AlignedFormatter(logging.Formatter):
        _WIDTH = len("[CRITICAL]")
        def format(self, record):
            bracket = f"[{record.levelname}]"
            record.bracket = f"{bracket:<{self._WIDTH}}"
            return super().format(record)

    fmt = AlignedFormatter(
        fmt="%(asctime)s %(bracket)s  %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)
    logger.addHandler(ch)

    log_dir = Path(__file__).parent / "logs"
    log_dir.mkdir(exist_ok=True)
    log_file = log_dir / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    logger.info(f"로그 파일: {log_file}")
    return logger

log = setup_logger()

# ── 의존성 ───────────────────────────────────
try:
    import duckdb
except ImportError:
    log.critical("pip install duckdb"); sys.exit(1)

try:
    import pandas as pd
except ImportError:
    log.critical("pip install pandas"); sys.exit(1)

try:
    from dat_loader import (read_fwf_dat, read_fwf_duckdb, read_pipe_dat, read_pipe_duckdb,
                             read_csv_file, read_csv_duckdb, read_sas7bdat_file)
except ImportError:
    log.critical("dat_loader.py 없음"); sys.exit(1)


# ══════════════════════════════════════════════
# 설정
# ══════════════════════════════════════════════
ROOT = Path(__file__).parent
MAX_READ_WORKERS = 4
LOAD_TIMEOUT = 300          # 테이블당 최대 읽기 시간 (초), 0이면 무제한
ENCODINGS = ["cp949", "utf-8"]
FILE_EXTENSIONS = [".zip", ".dat.gz", ".DAT", ".dat", ".prn", ".csv", ".csv.gz", ".sas7bdat"]

PREFIX_OUT = "out_"


# ══════════════════════════════════════════════
# 유틸 (JOB 파일에서도 import하여 사용)
# ══════════════════════════════════════════════

def _dw(s):
    """터미널 표시 너비 (한글 2칸)"""
    return sum(2 if unicodedata.east_asian_width(c) in ('W', 'F') else 1 for c in s)


def _pad(s, width):
    """한글 폭 고려 오른쪽 패딩 (ljust 대체)"""
    return s + ' ' * max(0, width - _dw(s))


def sql(con, label, query, params=None):
    """SQL 실행 + CREATE TABLE이면 건수 로깅"""
    t = time.time()
    con.execute(query, params or [])
    m = re.search(r"CREATE\s+(?:OR\s+REPLACE\s+)?TABLE\s+(?:IF\s+NOT\s+EXISTS\s+)?(\w+)",
                  query, re.IGNORECASE)
    if m:
        tbl = m.group(1)
        cnt = con.execute(f"SELECT COUNT(*) FROM {tbl}").fetchone()[0]
        log.info(f"  {_pad(label, 50)}  {cnt:>12,}건  ({time.time()-t:.1f}초)")


def prev_ym(yyyymm, n=1):
    """YYYYMM 기준 n개월 전 YYYYMM 문자열 반환 (기본 1개월 전)"""
    y, m = int(yyyymm[:4]), int(yyyymm[4:])
    m -= n
    while m < 1:
        y -= 1
        m += 12
    while m > 12:
        y += 1
        m -= 12
    return f"{y}{m:02d}"


def table_exists(con, name):
    """테이블 존재 여부"""
    cnt = con.execute(
        f"SELECT COUNT(*) FROM information_schema.tables "
        f"WHERE table_schema='main' AND table_name='{name}'"
    ).fetchone()[0]
    return cnt > 0


def require_tables(con, *names):
    """모든 테이블이 존재하는지 확인. 없는 테이블은 경고 로그 출력."""
    missing = [n for n in names if not table_exists(con, n)]
    if missing:
        for n in missing:
            log.warning(f"  [--] {_pad(n, 45)}  테이블 없음 — 건너뜀")
        return False
    return True


def check(con, label, query, expect="zero"):
    """
    검증용 헬퍼. SELECT 결과를 로깅만 하고 테이블은 만들지 않음.

    expect:
      "zero"    — 0건이어야 정상 (이상 데이터 탐지)
      "nonzero" — 1건 이상이어야 정상 (데이터 존재 확인)
      int       — 정확히 N건이어야 정상
    """
    try:
        row = con.execute(query).fetchone()
    except Exception as e:
        if "not found" in str(e).lower() or "does not exist" in str(e).lower():
            log.warning(f"  [--] {_pad(label, 45)}  테이블 없음 — 건너뜀")
            return False
        raise
    cnt = row[0] if row else 0
    if expect == "zero":
        ok = cnt == 0
    elif expect == "nonzero":
        ok = cnt > 0
    else:
        ok = cnt == expect
    mark = "OK" if ok else "NG"
    log.info(f"  [{mark}] {_pad(label, 45)}  {cnt:>12,}건")
    return ok


def check_sum(con, label, query):
    """
    합계 표시용 헬퍼. SELECT 결과(단일 행)의 컬럼들을 로깅.

    사용 예:
      check_sum(con, "AMT 합계", "SELECT SUM(AMT) AS AMT FROM tfc81")
      check_sum(con, "보험료 합계",
                "SELECT SUM(RP_PRM) AS RP_PRM, SUM(AF_PRM) AS AF_PRM FROM fio841")
    """
    try:
        row = con.execute(query).fetchone()
    except Exception as e:
        if "not found" in str(e).lower() or "does not exist" in str(e).lower():
            log.warning(f"  [--] {_pad(label, 45)}  테이블 없음 — 건너뜀")
            return
        raise
    cols = [d[0] for d in con.description]
    if not row:
        log.info(f"  [--] {_pad(label, 45)}  데이터 없음")
        return
    parts = []
    for c, v in zip(cols, row):
        if v is None:
            parts.append(f"{c}: NULL")
        elif isinstance(v, (int, float)):
            parts.append(f"{c}: {v:>14,.0f}")
        else:
            parts.append(f"{c}: {v}")
    log.info(f"  [OK] {_pad(label, 45)}  {' | '.join(parts)}")


def check_diff(con, label, query_a, query_b, group_cols, sum_col,
               threshold=1):
    """
    두 쿼리 결과를 group_cols 기준 FULL OUTER JOIN하여
    sum_col 차이가 threshold 초과인 행을 로그 출력.

    Returns: bool — 차이 없으면 True
    """
    join_cond = " AND ".join(
        f"a.{c} IS NOT DISTINCT FROM b.{c}" for c in group_cols)
    coalesce_keys = ", ".join(
        f"COALESCE(a.{c}, b.{c}) AS {c}" for c in group_cols)

    diff_sql = f"""
        WITH a AS ({query_a}),
             b AS ({query_b})
        SELECT {coalesce_keys},
               COALESCE(a.{sum_col}, 0) AS val_a,
               COALESCE(b.{sum_col}, 0) AS val_b,
               COALESCE(a.{sum_col}, 0) - COALESCE(b.{sum_col}, 0) AS diff
        FROM a FULL OUTER JOIN b ON {join_cond}
        WHERE ABS(COALESCE(a.{sum_col}, 0) - COALESCE(b.{sum_col}, 0)) > {threshold}
        ORDER BY ABS(COALESCE(a.{sum_col}, 0) - COALESCE(b.{sum_col}, 0)) DESC
    """

    try:
        rows = con.execute(diff_sql).fetchall()
    except Exception as e:
        if "not found" in str(e).lower() or "does not exist" in str(e).lower():
            log.warning(f"  [--] {_pad(label, 45)}  테이블 없음 — 건너뜀")
            return True
        raise
    total = len(rows)
    ok = total == 0
    mark = "OK" if ok else "NG"
    log.info(f"  [{mark}] {_pad(label, 45)}  차이 {total:,}건")

    if not ok:
        show = rows[:20]
        for r in show:
            keys = ", ".join(f"{c}={r[i]}" for i, c in enumerate(group_cols))
            val_a, val_b, diff = r[len(group_cols)], r[len(group_cols)+1], r[len(group_cols)+2]
            log.info(f"       {_pad(keys, 30)}  A: {val_a:>14,.0f}  B: {val_b:>14,.0f}  diff: {diff:>14,.0f}")
        if total > 20:
            log.info(f"       ... 외 {total - 20:,}건")
        # 합계
        gc = len(group_cols)
        sum_a = sum(r[gc] for r in rows)
        sum_b = sum(r[gc + 1] for r in rows)
        sum_d = sum(r[gc + 2] for r in rows)
        log.info(f"       {_pad('[합계]', 30)}  A: {sum_a:>14,.0f}  B: {sum_b:>14,.0f}  diff: {sum_d:>14,.0f}")

        # CSV 저장
        import csv
        out_dir = ROOT / "output"
        out_dir.mkdir(parents=True, exist_ok=True)
        safe_name = label.replace(" ", "_").replace("/", "_")[:50]
        csv_path = out_dir / f"diff_{safe_name}.csv"
        cols = group_cols + ["val_a", "val_b", "diff"]
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(cols)
            w.writerows(rows)
        log.info(f"       → {csv_path}")
    return ok


def row_count(con, table, group_by=None, where=None):
    """테이블 건수 로깅. group_by 지정 시 그룹별, where 지정 시 조건부 건수 출력."""
    if not table_exists(con, table):
        log.warning(f"  [--] {_pad(table, 45)}  테이블 없음")
        return 0
    w = f" WHERE {where}" if where else ""
    if group_by:
        cols = group_by if isinstance(group_by, str) else ", ".join(group_by)
        rows = con.execute(
            f"SELECT {cols}, COUNT(*) AS cnt FROM {table}{w} GROUP BY {cols} ORDER BY {cols}"
        ).fetchall()
        total = 0
        for row in rows:
            key = " | ".join(str(v) for v in row[:-1])
            total += row[-1]
            log.info(f"  [OK] {_pad(table, 30)}  {key:>14s}  {row[-1]:>12,}건")
        log.info(f"  [OK] {_pad(table, 30)}  {'합계':>14s}  {total:>12,}건")
        return total
    cnt = con.execute(f"SELECT COUNT(*) FROM {table}{w}").fetchone()[0]
    log.info(f"  [OK] {_pad(table, 45)}  {cnt:>12,}건")
    return cnt


def _resolve_path(base, file_template, yyyymm):
    """파일 확장자 순서대로 탐색"""
    stem = Path(file_template.format(yyyymm=yyyymm)).stem
    if stem.lower().endswith(".dat"):
        stem = stem[:-4]
    for ext in FILE_EXTENSIONS:
        candidate = base / f"{stem}{ext}"
        if candidate.exists():
            return candidate
    raise FileNotFoundError(f"파일 없음: {stem}.* (위치: {base})")


def _make_oracle_dsn(cfg):
    """Oracle DSN 생성. SERVICE_NAME(/) 과 SID(:) 모두 지원.

    dsn 형식:
      "host:port/service_name"  → SERVICE_NAME 방식
      "host:port:sid"           → SID 방식 (makedsn 사용)
      "host:port"  + sid 키     → SID 방식
      "host:port"  + service_name 키 → SERVICE_NAME 방식
    """
    import oracledb

    dsn = cfg.get("dsn", "")

    # 명시적 sid / service_name 키가 있으면 우선
    if "sid" in cfg:
        host_port = dsn.split("/")[0]  # host:port 부분만
        parts = host_port.split(":")
        return oracledb.makedsn(parts[0], parts[1] if len(parts) > 1 else "1521",
                                sid=cfg["sid"])
    if "service_name" in cfg:
        host_port = dsn.split("/")[0]
        parts = host_port.split(":")
        return oracledb.makedsn(parts[0], parts[1] if len(parts) > 1 else "1521",
                                service_name=cfg["service_name"])

    # dsn 문자열에서 자동 판별
    # host:port/service → SERVICE_NAME
    if "/" in dsn:
        return dsn

    # host:port:sid → SID
    parts = dsn.split(":")
    if len(parts) == 3:
        return oracledb.makedsn(parts[0], parts[1], sid=parts[2])

    return dsn


def _load_oracle(cfg, name, yyyymm):
    """Oracle DB에서 SQL로 데이터 추출 → DataFrame"""
    try:
        import oracledb
    except ImportError:
        raise ImportError(
            "oracledb 패키지가 필요합니다: pip install oracledb")

    dsn = _make_oracle_dsn(cfg)
    user = cfg.get("user", "")
    password = cfg.get("password", "")
    sql_text = cfg["sql"].replace("{yyyymm}", yyyymm)

    log.info(f"  [Oracle] {_pad(name, 20)} ← {dsn}")
    log.debug(f"  [Oracle] SQL: {sql_text[:120]}...")

    fetch_size = cfg.get("fetch_size", 50_000)

    try:
        with oracledb.connect(user=user, password=password, dsn=dsn) as conn:
            cur = conn.cursor()
            cur.arraysize = fetch_size
            cur.execute(sql_text)
            cols = [c[0] for c in cur.description]

            chunks = []
            while True:
                rows = cur.fetchmany(fetch_size)
                if not rows:
                    break
                chunks.append(rows)
                log.debug(f"  [Oracle] {_pad(name, 20)} {sum(len(c) for c in chunks):,}건 읽는 중...")

            import pandas as pd
            if chunks:
                import itertools
                df = pd.DataFrame(itertools.chain.from_iterable(chunks), columns=cols)
            else:
                df = pd.DataFrame(columns=cols)
    except Exception as e:
        log.error(f"  [Oracle] {name} 실패: {e}")
        log.error(f"  [Oracle] SQL:\n{sql_text}")
        raise

    log.info(f"  [Oracle] {_pad(name, 20)} {len(df):>12,}건")
    return df


def _load_file(path, cfg, name):
    """단일 파일 읽기 → DataFrame"""
    numeric = cfg.get("numeric", [])
    if cfg["type"] == "fwf":
        return read_fwf_dat(path, cfg["cols"], numeric=numeric,
                            encoding=cfg.get("encoding", "cp949"))
    elif cfg["type"] == "pipe":
        return read_pipe_dat(path, cfg["cols"], numeric=numeric,
                             encodings=ENCODINGS, delimiter=cfg.get("delimiter", "|"))
    elif cfg["type"] == "csv":
        return read_csv_file(path, cfg.get("cols"), numeric=numeric,
                             encodings=ENCODINGS,
                             delimiter=cfg.get("delimiter", ","),
                             header=cfg.get("header", False))
    elif cfg["type"] == "sas7bdat":
        return read_sas7bdat_file(path, numeric=numeric,
                                  encoding=cfg.get("encoding", "cp949"))
    raise ValueError(f"Unknown type: {cfg['type']}")


def _upsert(con, name, df, yyyymm, month_col):
    """월별 누적 적재: 해당 월 DELETE → INSERT"""
    tmp = f"_df_{name}"
    con.register(tmp, df)

    if not table_exists(con, name):
        con.execute(f"CREATE TABLE {name} AS SELECT * FROM {tmp}")
        return len(df)

    if month_col:
        con.execute(f"DELETE FROM {name} WHERE CAST({month_col} AS VARCHAR) LIKE '{yyyymm}%'")
    else:
        log.warning(f"  [{name}] month_col=null → 전체 교체 (기존 데이터 삭제)")
        con.execute(f"DELETE FROM {name}")

    con.execute(f"INSERT INTO {name} SELECT * FROM {tmp}")
    return len(df)


# ══════════════════════════════════════════════
# JOB 모듈 로더
# ══════════════════════════════════════════════

def load_job_module(job_path: str):
    """
    JOB 파일(.py)을 동적 로드.

    JOB 파일 필수 정의:
      NAME          : str   — JOB 이름 (예: "job1")
      TABLES        : dict  — 테이블 정의 {name: {type, file, cols, ...}}
      logic(con, yyyymm)    — 비즈니스 로직
      validate(con, yyyymm) — 검증 로직

    선택 정의:
      DESC          : str   — 설명
      EXPORT_SHEETS : dict  — {table_name: sheet_name}
    """
    path = Path(job_path).resolve()
    if not path.exists():
        raise FileNotFoundError(f"JOB 파일 없음: {path}")

    spec = importlib.util.spec_from_file_location(path.stem, path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)

    for attr in ("NAME", "TABLES", "logic", "validate"):
        if not hasattr(mod, attr):
            raise AttributeError(f"JOB 파일에 '{attr}' 없음: {path}")

    return mod


# ══════════════════════════════════════════════
# 파이프라인 (load → logic → validate → export)
# ══════════════════════════════════════════════

def _read_one(name, cfg, base_path, yyyymm):
    """단일 테이블 읽기 (병렬 실행용) — 파일/Oracle 모두 지원"""
    if "month_col" not in cfg:
        raise KeyError(
            f"[{name}] TABLES 정의에 'month_col' 필수 "
            f"(누적: 컬럼명 / 전체교체: null)")

    ts = time.time()
    if cfg.get("type") == "oracle":
        df = _load_oracle(cfg, name, yyyymm)
        log.info(f"  [Read] {_pad(name, 20)} {len(df):>12,}건  ({time.time()-ts:.1f}초)")
    else:
        path = _resolve_path(base_path, cfg["file"], yyyymm)
        log.info(f"  [Read] {_pad(name, 20)} ← {path.name}")
        df = _load_file(path, cfg, name)
        log.info(f"  [Read] {_pad(name, 20)} {len(df):>12,}건  ({time.time()-ts:.1f}초)")
    return name, cfg, df


def do_load(con, yyyymm, tables: dict, timeout: int = None):
    """TABLES dict 기반 로드 (FWF → DuckDB 네이티브, 나머지 → 병렬 읽기+순차 적재)

    timeout: 테이블당 최대 읽기 시간(초). None이면 LOAD_TIMEOUT 사용, 0이면 무제한.
    """
    base_path = ROOT / "data" / yyyymm
    loaded, failed = [], []
    tmo = timeout if timeout is not None else LOAD_TIMEOUT

    # 테이블 키의 {yyyymm} 플레이스홀더를 실제 월로 치환
    tables = {k.replace("{yyyymm}", yyyymm): v for k, v in tables.items()}

    # DuckDB 네이티브 대상 (pipe, csv) / 나머지 (fwf, sas7bdat, oracle 등) 분리
    # fwf 는 SAS colspec 이 바이트 위치이므로 SUBSTR(글자 기반) 대신
    # 바이트 슬라이싱 pandas 경로를 사용해야 cp949 한글 위치가 밀리지 않음
    # 단, 한글 없는 fwf 파일은 "native": True 를 명시하면 DuckDB 네이티브로 처리
    native_types = {"pipe", "csv"}
    def _is_native(cfg):
        if cfg.get("type") == "fwf" and cfg.get("native"):
            return True
        return cfg.get("type") in native_types
    native_tables = {k: v for k, v in tables.items() if _is_native(v)}
    other_tables = {k: v for k, v in tables.items() if k not in native_tables}

    # ── FWF / Pipe: DuckDB 네이티브 읽기 (pandas 우회) ──
    for name, cfg in native_tables.items():
        ttype = cfg["type"]
        timer = None
        ts = time.time()
        t = cfg.get("timeout", tmo)  # 테이블별 timeout 우선
        try:
            path = _resolve_path(base_path, cfg["file"], yyyymm)
            log.info(f"  [Read] {_pad(name, 20)} ← {path.name}")

            # 타임아웃 설정: 시간 초과 시 con.interrupt()로 쿼리 취소
            if t > 0:
                timer = threading.Timer(t, lambda: con.interrupt())
                timer.start()

            # 테이블 정의에 encoding 있으면 우선 사용
            enc_override = cfg.get("encoding")
            encs = [enc_override] if enc_override else None

            if ttype == "fwf":
                cnt = read_fwf_duckdb(con, path, cfg["cols"], cfg.get("numeric"),
                                      encodings=encs)
                tmp_table = "_fwf_parsed"
            elif ttype == "csv":
                cnt = read_csv_duckdb(con, path, cfg.get("cols"),
                                      cfg.get("numeric"),
                                      cfg.get("delimiter", ","),
                                      cfg.get("header", False),
                                      encodings=encs)
                tmp_table = "_csv_parsed"
            else:  # pipe
                cnt = read_pipe_duckdb(con, path, cfg["cols"],
                                       cfg.get("numeric"),
                                       cfg.get("delimiter", "|"),
                                       encodings=encs)
                tmp_table = "_pipe_parsed"

            if timer:
                timer.cancel()

            # tmp → 실제 테이블로 upsert
            month_col = cfg.get("month_col")
            if not table_exists(con, name):
                con.execute(f"CREATE TABLE {name} AS SELECT * FROM {tmp_table}")
            else:
                if month_col:
                    con.execute(f"DELETE FROM {name} WHERE CAST({month_col} AS VARCHAR) LIKE '{yyyymm}%'")
                else:
                    con.execute(f"DELETE FROM {name}")
                con.execute(f"INSERT INTO {name} SELECT * FROM {tmp_table}")
            con.execute(f"DROP TABLE IF EXISTS {tmp_table}")
            log.info(f"  [Read] {_pad(name, 20)} {cnt:>12,}건  ({time.time()-ts:.1f}초)")
            loaded.append(name)
        except Exception as e:
            if timer:
                timer.cancel()
            elapsed = time.time() - ts
            # 타임아웃 여부 판별
            if t > 0 and elapsed >= t - 1:
                log.error(f"  [Read] {_pad(name, 20)} 타임아웃 ({t}초 초과) — 건너뜀")
                failed.append(name)
                continue
            log.warning(f"  [Read] {_pad(name, 20)} DuckDB 실패({e}) → pandas 폴백")
            try:
                ts_fb = time.time()
                # pandas 폴백에도 타임아웃 적용
                pd_encs = [enc_override] if enc_override else ENCODINGS
                def _pandas_fallback():
                    if ttype == "fwf":
                        return read_fwf_dat(path, cfg["cols"],
                                            numeric=cfg.get("numeric"),
                                            encoding=cfg.get("encoding", "cp949"))
                    else:
                        return read_pipe_dat(path, cfg["cols"],
                                             numeric=cfg.get("numeric"),
                                             encodings=pd_encs,
                                             delimiter=cfg.get("delimiter", "|"))

                if t > 0:
                    with ThreadPoolExecutor(max_workers=1) as fb_pool:
                        fb_fut = fb_pool.submit(_pandas_fallback)
                        df = fb_fut.result(timeout=t)
                else:
                    df = _pandas_fallback()

                cnt = _upsert(con, name, df, yyyymm, cfg.get("month_col"))
                log.info(f"  [Read] {_pad(name, 20)} {cnt:>12,}건  ({time.time()-ts_fb:.1f}초)  (폴백)")
                loaded.append(name)
            except TimeoutError:
                log.error(f"  [Read] {_pad(name, 20)} 폴백 타임아웃 ({t}초 초과) — 건너뜀")
                failed.append(name)
            except Exception as e2:
                log.error(f"  [Read] {_pad(name, 20)} 폴백도 실패: {e2}")
                failed.append(name)

    # ── 나머지 (sas7bdat, oracle 등): 병렬 읽기 + 순차 적재 ──
    if other_tables:
        results = []
        workers = min(MAX_READ_WORKERS, len(other_tables))
        log.info(f"  병렬 Read 시작 (workers={workers}, 테이블={len(other_tables)}개)")
        with ThreadPoolExecutor(max_workers=workers) as pool:
            futures = {
                name: pool.submit(_read_one, name, cfg, base_path, yyyymm)
                for name, cfg in other_tables.items()
            }
            for name, fut in futures.items():
                t = other_tables[name].get("timeout", tmo)  # 테이블별 timeout 우선
                try:
                    result_timeout = t if t > 0 else None
                    results.append(fut.result(timeout=result_timeout))
                except TimeoutError:
                    log.error(f"  [Read] {_pad(name, 20)} 타임아웃 ({t}초 초과) — 건너뜀")
                    failed.append(name)
                except FileNotFoundError:
                    log.warning(f"  [Read] {_pad(name, 20)} 파일 없음 — 건너뜀")
                    failed.append(name)
                except Exception as e:
                    log.error(f"  [Read] {_pad(name, 20)} 실패: {e}")
                    failed.append(name)

        log.info(f"  ── Load 시작 ({len(results)}개 테이블) ──")
        for name, cfg, df in results:
            try:
                ts = time.time()
                cnt = _upsert(con, name, df, yyyymm, cfg.get("month_col"))
                log.info(f"  [Load] {_pad(name, 20)} {cnt:>12,}건  ({time.time()-ts:.1f}초)")
                loaded.append(name)
            except Exception as e:
                log.error(f"  [Load] {_pad(name, 20)} 실패: {e}")
                failed.append(name)

    return loaded, failed


def _build_export_query(tbl, cfg, yyyymm=None):
    """EXPORT_SHEETS 값(str 또는 dict)으로부터 SQL과 시트명을 생성한다.

    지원 형식:
      문자열  → "시트명"                    → SELECT * FROM tbl
      dict   → {"sheet": "시트명"}          → SELECT * FROM tbl
             → {"sheet": .., "sql": "..."}  → 사용자 SQL 그대로
             → {"sheet": .., "columns": []} → SELECT col1,col2 FROM tbl
             → {"sheet": .., "where": ".."}→ SELECT * FROM tbl WHERE ..
             → columns + where 조합도 가능
    """
    if isinstance(cfg, str):
        sheet = cfg.replace("{yyyymm}", yyyymm) if yyyymm else cfg
        return f"SELECT * FROM {tbl}", sheet

    sheet = cfg.get("sheet", tbl)
    if yyyymm:
        sheet = sheet.replace("{yyyymm}", yyyymm)

    # sql이 있으면 그대로 사용 (columns/where 무시)
    if "sql" in cfg:
        query = cfg["sql"].replace("{yyyymm}", yyyymm) if yyyymm else cfg["sql"]
        return query, sheet

    cols = ", ".join(cfg["columns"]) if "columns" in cfg else "*"
    query = f"SELECT {cols} FROM {tbl}"
    if "where" in cfg:
        query += f" WHERE {cfg['where']}"
    if "order_by" in cfg:
        query += f" ORDER BY {cfg['order_by']}"
    if "limit" in cfg:
        query += f" LIMIT {cfg['limit']}"

    if yyyymm:
        query = query.replace("{yyyymm}", yyyymm)
    return query, sheet


def _next_output_path(out_dir, job_name, yyyymm):
    """출력 파일 경로 결정. 항상 다음 마이너 순번으로 생성.

    v0.01 → v0.02 → ... → v0.99 → v0.100 → ...
    유저가 v1.0 또는 v1.00을 만들면 → v1.01 → v1.02 → ...
    메이저 번호는 유저가 직접 올린다.
    """
    import glob as _glob
    import re
    prefix = f"{job_name}_{yyyymm}"

    # 기존 파일에서 메이저별 최대 마이너 찾기
    # v1.0, v1.00 둘 다 마이너=0 으로 인식
    existing = _glob.glob(str(out_dir / f"{prefix}_v*.xlsx"))
    max_major = 0
    max_minor = 0  # 해당 메이저의 최대 마이너
    for f in existing:
        fname = Path(f).stem
        # v1.04, v1.0, v1.00 → major.minor / v1, v2 → major only (minor=0)
        m = re.search(r"_v(\d+)(?:\.(\d+))?$", fname)
        if not m:
            continue
        major = int(m.group(1))
        minor = int(m.group(2)) if m.group(2) is not None else 0
        if major > max_major:
            max_major = major
            max_minor = minor
        elif major == max_major:
            max_minor = max(max_minor, minor)

    if not existing:
        return out_dir / f"{prefix}_v0.01.xlsx"

    next_minor = max_minor + 1
    return out_dir / f"{prefix}_v{max_major}.{next_minor:02d}.xlsx"


def do_export(con, yyyymm, job_name, sheet_map):
    """시트맵 기반 Excel 출력"""
    out_dir = ROOT / "output"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = _next_output_path(out_dir, job_name, yyyymm)

    # 테이블 키의 {yyyymm} 플레이스홀더를 실제 월로 치환
    sheet_map = {k.replace("{yyyymm}", yyyymm): v for k, v in sheet_map.items()}

    # out_ 접두사 테이블 자동 추가
    db_tables = [r[0] for r in con.execute(
        "SELECT table_name FROM information_schema.tables "
        "WHERE table_schema='main' ORDER BY table_name"
    ).fetchall()]
    for tbl in db_tables:
        if tbl.startswith(PREFIX_OUT) and tbl not in sheet_map:
            sheet_map[tbl] = tbl[len(PREFIX_OUT):]

    summary = []
    with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
        for tbl, cfg in sheet_map.items():
            # 리스트이면 같은 테이블에서 여러 시트 추출
            cfgs = cfg if isinstance(cfg, list) else [cfg]
            for single_cfg in cfgs:
                try:
                    ts = time.time()
                    sql, sheet = _build_export_query(tbl, single_cfg, yyyymm)
                    df = con.execute(sql).df()
                    sname = sheet[:31]
                    df.to_excel(writer, sheet_name=sname, index=False)
                    # 정수/실수 컬럼에 #,##0 서식 적용 (E+ 방지)
                    ws = writer.sheets[sname]
                    for col_idx, dtype in enumerate(df.dtypes, 1):
                        if dtype.kind in ('i', 'u', 'f'):
                            fmt = '#,##0.##' if dtype.kind == 'f' else '#,##0'
                            for row_idx in range(2, len(df) + 2):
                                ws.cell(row=row_idx, column=col_idx).number_format = fmt
                    el = time.time() - ts
                    summary.append((tbl, sname, len(df), el))
                    log.info(f"    {_pad(sname, 25)}  {len(df):>10,}건  ({el:.1f}초)")
                except Exception as e:
                    err_sheet = (single_cfg if isinstance(single_cfg, str) else single_cfg.get("sheet", tbl)).replace("{yyyymm}", yyyymm)
                    log.warning(f"    {err_sheet} 건너뜀: {e}")

        _write_summary_sheet(writer, yyyymm, summary)

    log.info(f"  → {out_file}")
    return out_file


def _write_summary_sheet(writer, yyyymm, summary):
    """요약 시트 생성"""
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = writer.book
    ws = wb.create_sheet("요약", 0)

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    alt_fill = PatternFill("solid", fgColor="EBF3FB")
    link_font = Font(color="1F4E79", underline="single")
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")
    right  = Alignment(horizontal="right",  vertical="center")
    thin   = Side(style="thin", color="BFBFBF")
    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = f"출력 요약  |  {yyyymm}  |  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    c.font, c.alignment = Font(bold=True, size=13, color="1F4E79"), left
    ws.row_dimensions[1].height = 28

    for col, (h, w) in enumerate(zip(
        ["No", "테이블명", "시트명", "건수", "소요(초)"], [6, 28, 28, 16, 10]
    ), 1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill, c.font, c.alignment, c.border = hdr_fill, hdr_font, center, bdr
        ws.column_dimensions[get_column_letter(col)].width = w

    for i, (tbl, sname, cnt, el) in enumerate(summary, 1):
        row = i + 2
        fill = alt_fill if i % 2 == 0 else PatternFill()
        ws.cell(row=row, column=1, value=i).alignment = center
        ws.cell(row=row, column=2, value=tbl).alignment = left
        c_s = ws.cell(row=row, column=3, value=sname)
        c_s.hyperlink, c_s.font, c_s.alignment = f"#'{sname}'!A1", link_font, left
        c_c = ws.cell(row=row, column=4, value=cnt)
        c_c.number_format, c_c.alignment = "#,##0", right
        ws.cell(row=row, column=5, value=round(el, 1)).alignment = center
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = bdr
            ws.cell(row=row, column=col).fill = fill

    sr = len(summary) + 3
    total = sum(r[2] for r in summary)
    ws.merge_cells(f"A{sr}:C{sr}")
    c_l = ws.cell(row=sr, column=1, value="합계")
    c_l.font, c_l.alignment, c_l.border = Font(bold=True), center, bdr
    c_t = ws.cell(row=sr, column=4, value=total)
    c_t.number_format, c_t.font, c_t.alignment, c_t.border = "#,##0", Font(bold=True), right, bdr


def run_job(con, job_mod, yyyymm, skip_load=False, stages=None,
            only_tables=None, load_timeout=None):
    """단일 JOB 실행: load → logic → validate → export"""
    name = job_mod.NAME
    desc = getattr(job_mod, "DESC", "")
    # stages가 지정되면 해당 단계만 실행, 아니면 전체 실행
    run_all = stages is None

    log.info("=" * 60)
    log.info(f"[{name}] {desc}  (월: {yyyymm})")
    if not run_all:
        log.info(f"[{name}] 선택 단계: {stages}")
    if only_tables:
        log.info(f"[{name}] 선택 테이블: {only_tables}")
    log.info("=" * 60)
    t0 = time.time()

    # 1. LOAD
    if run_all and skip_load:
        log.info(f"[{name}] LOAD 스킵")
    elif run_all or "load" in stages:
        tables = job_mod.TABLES
        if only_tables:
            unknown = set(only_tables) - set(tables)
            if unknown:
                log.warning(f"[{name}] 존재하지 않는 테이블 무시: {unknown}")
            tables = {k: v for k, v in tables.items() if k in only_tables}
            if not tables:
                log.warning(f"[{name}] 로드할 테이블이 없습니다")
        tmo_label = f", 타임아웃={load_timeout}초" if load_timeout else ""
        log.info(f"[{name}] LOAD ({len(tables)}개 테이블{tmo_label})")
        loaded, failed = do_load(con, yyyymm, tables, timeout=load_timeout)
        if failed:
            log.warning(f"[{name}] 로드 실패: {failed}")

    # 2. LOGIC
    if run_all or "logic" in stages:
        try:
            log.info(f"[{name}] LOGIC")
            job_mod.logic(con, yyyymm)
        except Exception as e:
            log.error(f"[{name}] LOGIC 실패 — 다음 단계로 계속 진행: {e}")

    # 3. VALIDATE
    if run_all or "validate" in stages:
        try:
            log.info("─" * 60)
            log.info(f"[{name}] VALIDATE")
            job_mod.validate(con, yyyymm)
            log.info("─" * 60)
        except Exception as e:
            log.error(f"[{name}] VALIDATE 실패 — 다음 단계로 계속 진행: {e}")

    # 4. EXPORT
    if run_all or "export" in stages:
        try:
            sheets = dict(getattr(job_mod, "EXPORT_SHEETS", {}))
            if sheets:
                log.info(f"[{name}] EXPORT")
                do_export(con, yyyymm, name, sheets)
        except Exception as e:
            log.error(f"[{name}] EXPORT 실패: {e}")

    log.info(f"[{name}] 완료  소요: {time.time()-t0:.1f}초")


# ══════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="SAS → DuckDB 사업비배분 파이프라인",
        formatter_class=argparse.RawTextHelpFormatter,
        epilog="""
예시:
  python sas_to_duckdb.py --ym 202601 --job jobs/job1.py
  python sas_to_duckdb.py --ym 202601 --job jobs/job2.py --skip-load
  python sas_to_duckdb.py --ym 202601 --job jobs/job1.py jobs/job2.py jobs/job3.py
  python sas_to_duckdb.py --ym 202601 202602 --job jobs/job1.py jobs/job2.py  # 월별 순차
  python sas_to_duckdb.py --ym 202601 --job jobs/job1.py --stage load
  python sas_to_duckdb.py --ym 202601 --job jobs/job1.py -s logic validate
  python sas_to_duckdb.py --ym 202601 --job jobs/job1.py -t fio841 fio842
        """
    )
    parser.add_argument("--ym", type=str, nargs="+", default=["202601"],
                        help="처리할 년월 (복수 가능, 예: --ym 202601 202602)")
    parser.add_argument("--job", "-j", nargs="+", required=True,
                        help="실행할 JOB 파일 경로 (예: jobs/job1.py)")
    parser.add_argument("--skip-load", action="store_true",
                        help="LOAD 단계 생략")
    parser.add_argument("--stage", "-s", nargs="+",
                        choices=["load", "logic", "validate", "export"],
                        help="실행할 단계만 지정 (예: --stage load)")
    parser.add_argument("--tables", "-t", nargs="+",
                        help="로드할 테이블명만 지정 (예: --tables fio841 fio842)")
    parser.add_argument("--load-timeout", type=int, default=None,
                        help=f"테이블당 최대 읽기 시간(초) (기본: {LOAD_TIMEOUT}, 0=무제한)")
    parser.add_argument("--verbose", "-v", action="store_true",
                        help="DEBUG 로그 콘솔 출력")

    args = parser.parse_args()

    log.info(f"Python {sys.version.split()[0]}")
    log.info(f"duckdb {duckdb.__version__}")
    log.info(f"pandas {pd.__version__}")

    if args.verbose:
        for h in log.handlers:
            if isinstance(h, logging.StreamHandler) and not isinstance(h, logging.FileHandler):
                h.setLevel(logging.DEBUG)

    ym_list = args.ym
    db_path = ROOT / "db" / "ifrs4-expense.duckdb"
    db_path.parent.mkdir(parents=True, exist_ok=True)
    log.info(f"대상 월: {', '.join(ym_list)}")
    log.info(f"DB    : {db_path}")

    # JOB 모듈 로드
    job_mods = []
    for jp in args.job:
        try:
            job_mods.append(load_job_module(jp))
        except (FileNotFoundError, AttributeError) as e:
            log.error(str(e))
            return

    # 순차 실행: ym → job 순서
    con = duckdb.connect(str(db_path))
    try:
        t_total = time.time()
        failed_jobs = []
        for yyyymm in ym_list:
            if len(ym_list) > 1:
                log.info("═" * 60)
                log.info(f"▶ 월별 실행 시작: {yyyymm}")
                log.info("═" * 60)
            for mod in job_mods:
                try:
                    run_job(con, mod, yyyymm, skip_load=args.skip_load,
                            stages=args.stage, only_tables=args.tables,
                            load_timeout=args.load_timeout)
                except Exception as e:
                    name = getattr(mod, "NAME", str(mod))
                    log.error(f"[{name}] 실패 — 다음 JOB으로 계속 진행: {e}")
                    failed_jobs.append(f"{yyyymm}/{name}")
        if failed_jobs:
            log.warning(f"실패한 JOB: {failed_jobs}")
        log.info(f"전체 완료  총 소요: {time.time()-t_total:.1f}초")
    finally:
        con.close()


if __name__ == "__main__":
    main()
