"""
SAS → DuckDB 전환 파이프라인 (단순화 버전)

실행 예시:
  python "sas_to_duckdb 12.py" --ym 202601 --stage load logic validate export
  python "sas_to_duckdb 12.py" --ym 202601 --stage load --tables erp sa01
  python "sas_to_duckdb 12.py" --ym 202601 --stage logic validate export
  python "sas_to_duckdb 12.py" --list
"""

import re
import sys
import logging
import time
import argparse
import unicodedata
from pathlib import Path
from datetime import datetime

# ══════════════════════════════════════════════
# 로거 설정
# ══════════════════════════════════════════════
def setup_logger() -> logging.Logger:
    logger = logging.getLogger("pipeline")
    logger.setLevel(logging.DEBUG)

    class AlignedFormatter(logging.Formatter):
        _WIDTH = len("[CRITICAL]")
        def format(self, record):
            bracket = f"[{record.levelname}]"
            record.bracket = f"{bracket:<{self._WIDTH}}"
            return super().format(record)

    fmt = AlignedFormatter(
        fmt     = "%(asctime)s %(bracket)s  %(message)s",
        datefmt = "%Y-%m-%d %H:%M:%S",
    )

    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)
    logger.addHandler(ch)

    log_dir = Path(__file__).parent / "logs"
    log_dir.mkdir(exist_ok=True)
    log_file = log_dir / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    logger.info(f"로그 파일: {log_file}")
    return logger

log = setup_logger()

# ── 의존성 체크 ───────────────────────────────
log.info(f"Python {sys.version.split()[0]}")

try:
    import duckdb
    log.info(f"duckdb {duckdb.__version__} 로드 OK")
except ImportError:
    log.critical("duckdb 없음 → pip install duckdb")
    sys.exit(1)

try:
    import pandas as pd
    log.info(f"pandas {pd.__version__} 로드 OK")
except ImportError:
    log.critical("pandas 없음 → pip install pandas")
    sys.exit(1)

try:
    from dat_loader import read_fwf_dat, read_pipe_dat, read_sas7bdat_file
    log.info("dat_loader.py 로드 OK")
except ImportError:
    log.critical("dat_loader.py 없음 → sas_to_duckdb.py 와 같은 폴더에 있어야 함")
    sys.exit(1)


# ══════════════════════════════════════════════
# 0. 설정
# ══════════════════════════════════════════════
ROOT      = Path(__file__).parent
ENCODINGS = ["cp949", "utf-8"]

# 파일 확장자 탐색 순서
FILE_EXTENSIONS = [".zip", ".dat.gz", ".DAT", ".dat", ".prn", ".csv", ".csv.gz", ".sas7bdat"]

# 테이블 네임스페이스 규칙
PREFIX_VALIDATE = "val_"   # 검증용 임시 → validate 후 자동 DROP
PREFIX_EXPORT   = "out_"   # Excel 출력 대상 → 시트명 = out_ 뗀 이름


# ══════════════════════════════════════════════
# 1. 테이블 레지스트리
# ══════════════════════════════════════════════
TABLE_REGISTRY = {

    # ── 유형A: 고정폭 ──────────────────────────────────────────────
    "fio841": {
        "type": "fwf",
        "file": "fioBtLtrJ841_01_{YYYYMM}.DAT",
        "desc": "수입보험료(유지비·신계약비 원천)",
        "cols": [
            ((0,    6),  "CLS_YYMM"),
            ((16,  32),  "PLYNO"),
            ((32,  37),  "INCM_PRM_CR_SEQNO"),
            ((37,  45),  "CNRDT"),
            ((45,  55),  "GDCD"),
            ((55,  65),  "IKD_GRPCD"),
            ((65,  75),  "INS_IMCD"),
            ((75,  85),  "CVRCD"),
            ((85,  95),  "FNL_CR_STCD"),
            ((95, 105),  "FNL_CR_DT_STCD"),
            ((105,113),  "FNL_CR_ST_CHDT"),
            ((113,123),  "NW_RNW_FLGCD"),
            ((123,128),  "PYM_SEQ"),
            ((128,134),  "FNL_PYM_YYMM"),
            ((134,149),  "RP_PRM"),
            ((149,164),  "AF_PRM"),
            ((164,174),  "DP_CASCD"),
            ((174,184),  "DP_DI_CASCD"),
            ((184,192),  "RV_CCLDT"),
            ((192,200),  "PPDT"),
            ((200,210),  "PYM_CYCCD"),
            ((210,217),  "DH_ORGCD"),
            ((217,224),  "DH_STFNO"),
            ((224,226),  "SL_TPCD"),
            ((226,241),  "CONDT_I_PRM"),
            ((241,251),  "UDRTK_TYCD"),
            ((251,252),  "MDF_CVAV"),
            ((252,253),  "MPY_CV_PRM"),
        ],
    },

    "fio843": {
        "type": "fwf",
        "file": "fioBtLtrJ843_01_{YYYYMM}.DAT",
        "desc": "사업비배분 보유건수",
        "cols": [
            ((0,    6),  "CLS_YYMM"),
            ((16,  32),  "PLYNO"),
            ((32,  40),  "CNRDT"),
            ((40,  50),  "GDCD"),
            ((50,  60),  "IKD_GRPCD"),
            ((60,  70),  "INS_IMCD"),
            ((70,  85),  "AP_PRM"),
            ((85,  92),  "DH_ORGCD"),
            ((92,  99),  "DH_STFNO"),
            ((99, 101),  "SL_TPCD"),
            ((101,111),  "UDRTK_TYCD"),
            ((111,116),  "BZCS_DV_PS"),
        ],
    },

    # ── 유형B: 파이프 구분자 ──────────────────────────────────────
    "erp": {
        "type": "pipe",
        "file": "BS100_{YYYYMM}.DAT",
        "desc": "ERP 전표 (BS100)",
        "cols": [
            "SLPDT",           # 전표일자
            "CLS_DA_SEQNO",    # 마감자료순번
            "NTACC_CD",        # 계정과목코드
            "DP_ORGCD",        # 발의기관코드
            "IKD_GRPCD",       # 보종군코드
            "INS_IMCD",        # 보험종목코드
            "RR_ORGCD",        # 귀속기관코드
            "SLP_AMT_DFRN_YN", # 전표금액차이여부
            "SLPNO",           # 전표번호
            "SLP_LNNO",        # 전표라인번호
            "BZCS_01_DVCD",    # 사업비1자배분코드
            "BZCS_02_DVCD",    # 사업비2자배분코드
            "FNDCD",           # 펀드코드
            "WONCR_POAMT",     # 원화계상금액
            "NOTS_MTT",        # 적요사항
            "BZCS_DV_EXEC_YN", # 사업비배분수행여부
            "INP_USR_ID",      # 입력사용자ID
            "INP_DTHMS",       # 입력일시
            "MDF_USR_ID",      # 수정사용자ID
            "MDF_DTHMS",       # 수정일시
        ],
    },

    "sa01": {
        "type": "pipe",
        "file": "RS100_{YYYYMM}.DAT",
        "desc": "유지비·신계약비 배분결과 (RS100)",
        "cols": [
            "SLPDT",       # 전표일자
            "SLPNO",       # 전표번호
            "SLP_LNNO",    # 전표라인번호
            "SNO",         # 일련번호
            "RR_ORGCD",    # 귀속기관코드
            "BZCS_TPCD",   # 사업비유형코드
            "INS_IMCD",    # 보험종목코드
            "SL_TPCD",     # 판매유형코드
            "CVRCD",       # 담보코드
            "DV_RT",       # 배분비율
            "DVAMT",       # 배분금액
            "INP_USR_ID",  # 입력사용자ID
            "INP_DTHMS",   # 입력일시
            "MDF_USR_ID",  # 수정사용자ID
            "MDF_DTHMS",   # 수정일시
            "NTACC_CD",    # 계정과목코드
            "SUB_SNO",     # 서브일련번호
        ],
    },

    "sa02": {
        "type": "pipe",
        "file": "RS101_{YYYYMM}.DAT",
        "desc": "손해조사비 배분결과 (RS101)",
        "cols": [
            "SLPDT",          # 전표일자
            "SLPNO",          # 전표번호
            "SLP_LNNO",       # 전표라인번호
            "SNO",            # 일련번호
            "RR_ORGCD",       # 귀속기관코드
            "BZCS_TPCD",      # 사업비유형코드
            "INS_IMCD",       # 보험종목코드
            "SL_TPCD",        # 판매유형코드
            "CVRCD",          # 담보코드
            "CR_CVRCD",       # 계약담보코드
            "DV_RT",          # 배분비율
            "DVAMT",          # 배분금액
            "BZCS_INS_IMCD",  # 사업비보험종목코드
            "INP_USR_ID",     # 입력사용자ID
            "INP_DTHMS",      # 입력일시
            "MDF_USR_ID",     # 수정사용자ID
            "MDF_DTHMS",      # 수정일시
            "NTACC_CD",       # 계정과목코드
        ],
    },

    "bs101": {
        "type": "pipe",
        "file": "BS101_{YYYYMM}.DAT",
        "desc": "원수실적마감 (BS101)",
        "cols": [
            "CLS_YYMM",          # 마감년월
            "CLS_DA_SEQNO",      # 마감자료순번
            "IKD_GRPCD",         # 보종군코드
            "INS_IMCD",          # 보험종목코드
            "PLYNO",             # 증권번호
            "INCM_PRM_CR_SEQNO", # 수입보험료발생순번
            "CNRDT",             # 계약일자
            "GDCD",              # 상품코드
            "FNL_CR_STCD",       # 최종계약상태코드
            "FNL_CR_DT_STCD",    # 최종계약세부상태코드
            "FNL_CR_ST_CHDT",    # 최종계약상태변경일자
            "NW_RNW_FLGCD",      # 신규갱신구분코드
            "CVRCD",             # 담보코드
            "DH_ORGCD",          # 취급기관코드
            "DH_STFNO",          # 취급직원번호
            "SL_TPCD",           # 판매유형코드
            "UDRTK_TYCD",        # 인수형태코드
            "RV_CCLDT",          # 수납취소일자
            "PYM_CYCCD",         # 납입주기코드
            "NANUM_TRI_YN",      # 나눔특약여부
            "PYM_SEQ",           # 납입회차
            "FNL_PYM_YYMM",      # 최종납입년월
            "RP_PRM",            # 영수보험료
            "AP_PRM",            # 적용보험료
            "DP_CASCD",          # 입금원인코드
            "DP_DT_CASCD",       # 입금세부원인코드
            "PPDT",              # 계상일자
            "CONDT_I_PRM",       # 공동인수총보험료
            "MDF_CVAV",          # 수정환산실적
            "MPY_CV_PRM",        # 출납환산보험료
            "RR_ORGCD",          # 귀속기관코드
            "INP_USR_ID",        # 입력사용자ID
            "INP_DTHMS",         # 입력일시
            "MDF_USR_ID",        # 수정사용자ID
            "MDF_DTHMS",         # 수정일시
        ],
    },

    "bs104": {
        "type": "pipe",
        "file": "BS104_{YYYYMM}.DAT",
        "desc": "지급보험금마감 (BS104)",
        "cols": [
            "CLS_YYMM",       # 마감년월
            "CLS_DA_SEQNO",   # 마감자료순번
            "IKD_GRPCD",      # 보종군코드
            "INS_IMCD",       # 보험종목코드
            "CVRCD",          # 담보코드
            "PLYNO",          # 증권번호
            "CNRDT",          # 계약일자
            "GDCD",           # 상품코드
            "FNL_CR_STCD",    # 최종계약상태코드
            "FNL_CR_DT_STCD", # 최종계약세부상태코드
            "FNL_CR_ST_CHDT", # 최종계약상태변경일자
            "DH_ORGCD",       # 취급기관코드
            "DH_STFNO",       # 취급직원번호
            "SL_TPCD",        # 판매유형코드
            "UDRTK_TYCD",     # 인수형태코드
            "AP_PRM",         # 적용보험료
            "RCP_YYMM",       # 수령년월
            "RCP_NV_SEQNO",   # 수령순번
            "IDM_CLM_CVRCD",  # 손해담보코드
            "EX_RCV_FLGCD",   # 예외수령구분코드
            "PY_IBAMT",       # 지급입력금액
            "BZCS_DV_CLMCT",  # 사업비배분청구건수
            "BZCS_DV_PYCT",   # 사업비배분지급건수
            "PPDT",           # 계상일자
            "RR_ORGCD",       # 귀속기관코드
            "INP_USR_ID",     # 입력사용자ID
            "INP_DTHMS",      # 입력일시
            "MDF_USR_ID",     # 수정사용자ID
            "MDF_DTHMS",      # 수정일시
            "NANUM_TRI_YN",   # 나눔특약여부
            "CR_CVRCD",       # 계약담보코드
        ],
    },

    "bs105": {
        "type": "pipe",
        "file": "BS105_{YYYYMM}.DAT",
        "desc": "지급준비금마감 (BS105)",
        "cols": [
            "CLS_YYMM",       # 마감년월
            "CLS_DA_SEQNO",   # 마감자료순번
            "IKD_GRPCD",      # 보종군코드
            "INS_IMCD",       # 보험종목코드
            "PLYNO",          # 증권번호
            "CNRDT",          # 계약일자
            "GDCD",           # 상품코드
            "FNL_CR_STCD",    # 최종계약상태코드
            "FNL_CR_DT_STCD", # 최종계약세부상태코드
            "FNL_CR_ST_CHDT", # 최종계약상태변경일자
            "CLM_CVRCD",      # 청구담보코드
            "DH_ORGCD",       # 취급기관코드
            "DH_STFNO",       # 취급직원번호
            "SL_TPCD",        # 판매유형코드
            "UDRTK_TYCD",     # 인수형태코드
            "AP_PRM",         # 적용보험료
            "ISAMT",          # 이사금액
            "PY_RFAMT",       # 지급참조금액
            "RR_ORGCD",       # 귀속기관코드
            "INP_USR_ID",     # 입력사용자ID
            "INP_DTHMS",      # 입력일시
            "MDF_USR_ID",     # 수정사용자ID
            "MDF_DTHMS",      # 수정일시
            "NANUM_TRI_YN",   # 나눔특약여부
            "CR_CVRCD",       # 계약담보코드
        ],
    },
}

# 숫자형 캐스팅 대상 (테이블별)
NUMERIC_COLS = {
    "fio841": ["INCM_PRM_CR_SEQNO", "PYM_SEQ", "RP_PRM", "AF_PRM",
               "CONDT_I_PRM", "MDF_CVAV", "MPY_CV_PRM"],
    "fio843": ["AP_PRM", "BZCS_DV_PS"],
    "erp":    ["CLS_DA_SEQNO", "WONCR_POAMT"],
    "sa01":   ["DV_RT", "DVAMT"],
    "sa02":   ["DV_RT", "DVAMT"],
    "bs101":  ["CLS_DA_SEQNO", "INCM_PRM_CR_SEQNO", "PYM_SEQ",
               "RP_PRM", "AP_PRM", "CONDT_I_PRM", "MDF_CVAV", "MPY_CV_PRM"],
    "bs104":  ["CLS_DA_SEQNO", "AP_PRM", "PY_IBAMT",
               "BZCS_DV_CLMCT", "BZCS_DV_PYCT"],
    "bs105":  ["CLS_DA_SEQNO", "AP_PRM", "ISAMT", "PY_RFAMT"],
}


# ══════════════════════════════════════════════
# 2. 유틸
# ══════════════════════════════════════════════

def _display_width(s):
    """한글 포함 문자열의 터미널 표시 너비"""
    return sum(2 if unicodedata.east_asian_width(c) in ('W', 'F') else 1 for c in s)


def _exec(con, label: str, sql: str):
    """SQL 실행 + CREATE TABLE이면 건수 로깅"""
    t = time.time()
    con.execute(sql)
    m = re.search(r"CREATE\s+OR\s+REPLACE\s+TABLE\s+(\w+)", sql, re.IGNORECASE)
    if m:
        tbl = m.group(1)
        cnt = con.execute(f"SELECT COUNT(*) FROM {tbl}").fetchone()[0]
        pad = 50
        label_padded = label + " " * max(0, pad - _display_width(label))
        log.info(f"  {label_padded}  {cnt:>12,}건  ({time.time()-t:.1f}초)")


def resolve_path(base: Path, file_template: str, yyyymm: str) -> Path:
    """파일 템플릿에서 확장자를 바꿔가며 실제 존재하는 파일 찾기"""
    stem = Path(file_template.format(YYYYMM=yyyymm)).stem
    if stem.lower().endswith(".dat"):
        stem = stem[:-4]

    for ext in FILE_EXTENSIONS:
        candidate = base / f"{stem}{ext}"
        if candidate.exists():
            log.debug(f"  발견: {candidate.name}")
            return candidate

    raise FileNotFoundError(f"파일 없음: {stem}.* (탐색 위치: {base})")


# ══════════════════════════════════════════════
# 3. LOAD — DAT 파일 → DuckDB 테이블
# ══════════════════════════════════════════════

def step_load(con, yyyymm: str, tables: list = None):
    """DAT 파일을 읽어 DuckDB 테이블로 적재 (CREATE OR REPLACE)"""
    base_path = ROOT / "data" / yyyymm
    target = {k: v for k, v in TABLE_REGISTRY.items()
              if not tables or k in tables}

    t0 = time.time()
    log.info("=" * 55)
    log.info(f"[LOAD] 시작  (월: {yyyymm}, 대상: {list(target.keys())})")
    log.info("=" * 55)

    loaded, failed = [], []

    for name, cfg in target.items():
        try:
            path = resolve_path(base_path, cfg["file"], yyyymm)
            log.info(f"[{name}] {cfg['desc']}  ← {path.name}")
            ts = time.time()

            if cfg["type"] == "fwf":
                df = read_fwf_dat(path, cfg["cols"],
                                  numeric=NUMERIC_COLS.get(name, []),
                                  encodings=ENCODINGS)
            elif cfg["type"] == "pipe":
                df = read_pipe_dat(path, cfg["cols"],
                                   numeric=NUMERIC_COLS.get(name, []),
                                   encodings=ENCODINGS,
                                   delimiter=cfg.get("delimiter", "|"))
            elif cfg["type"] == "sas7bdat":
                df = read_sas7bdat_file(path,
                                        numeric=NUMERIC_COLS.get(name, []),
                                        encoding=cfg.get("encoding", "cp949"))
            else:
                raise ValueError(f"Unknown type: {cfg['type']}")

            con.register(f"_df_{name}", df)
            con.execute(f"CREATE OR REPLACE TABLE {name} AS SELECT * FROM _df_{name}")
            cnt = len(df)
            log.info(f"[{name}] {cnt:,}건 적재 완료  ({time.time()-ts:.1f}초)")
            loaded.append(name)

        except FileNotFoundError as e:
            log.warning(f"[{name}] 파일 없음 — 건너뜀  ({e})")
            failed.append(name)
        except Exception as e:
            log.error(f"[{name}] 로드 실패: {type(e).__name__}: {e}")
            failed.append(name)

    log.info(f"[LOAD] 완료  성공: {loaded}  실패: {failed}  소요: {time.time()-t0:.1f}초")
    return loaded


# ══════════════════════════════════════════════
# 4. LOGIC — 비즈니스 로직 (SAS → SQL 변환)
# ══════════════════════════════════════════════

def step_logic(con, yyyymm: str):
    """배분 비즈니스 로직 실행"""
    t0 = time.time()
    log.info("=" * 55)
    log.info("[LOGIC] 배분 로직 실행 시작")
    log.info("=" * 55)

    # 필수 테이블 확인
    existing = {r[0] for r in con.execute(
        "SELECT table_name FROM information_schema.tables "
        "WHERE table_schema = 'main'"
    ).fetchall()}
    missing = {"fio841", "fio843", "sa01", "sa02"} - existing
    if missing:
        log.warning(f"필수 테이블 없음: {missing} — load 단계를 먼저 실행하세요")
        return

    # ── 841 보험료집계 ────────────────────────
    _exec(con, "BASE_DATA_CH (841 집계)", """
        CREATE OR REPLACE TABLE BASE_DATA_CH AS
        SELECT  INS_IMCD
               ,SUM(RP_PRM) AS RP_PRM
               ,SUM(AP_PRM) AS AP_PRM
        FROM fio841
        GROUP BY INS_IMCD
    """)

    # ── 843 보험료집계 ────────────────────────
    _exec(con, "BASE_DATA_CH_843 (843 집계)", """
        CREATE OR REPLACE TABLE BASE_DATA_CH_843 AS
        SELECT  INS_IMCD
               ,SUM(AP_PRM) AS AP_PRM
        FROM fio843
        GROUP BY INS_IMCD
    """)

    # ── 준비금 마감후 보유계약건수 ──
    _exec(con, "BASE_DATA_CH2 (843 건수집계)", """
        CREATE OR REPLACE TABLE BASE_DATA_CH2 AS
        SELECT  CLS_YYMM
               ,INS_IMCD
               ,SUM(AP_PRM)     AS AP_PRM
               ,SUM(BZCS_DV_PS) AS BZCS_DV_PS
        FROM fio843
        GROUP BY CLS_YYMM, INS_IMCD
    """)

    # ── 취급기관 이상 여부 확인 ──
    _exec(con, "CH (취급기관 이상체크)", """
        CREATE OR REPLACE TABLE CH AS
        SELECT DISTINCT DH_ORGCD, DH_STFNO
        FROM fio843
        WHERE SUBSTR(DH_STFNO, 1, 2) IN ('2S')
          AND SUBSTR(DH_ORGCD, 1, 1) NOT IN ('E')
    """)

    # ── SA11: sa01 변환 ───────────────────────
    _exec(con, "SA11 (sa01 변환)", """
        CREATE OR REPLACE TABLE sa11 AS
        SELECT
             SLPDT                                          AS YYYYMM
            ,NTACC_CD
            ,BZCS_TPCD
            ,RR_ORGCD
            ,SL_TPCD
            ,INS_IMCD
            ,CASE WHEN CVRCD = '*' THEN '0' ELSE CVRCD END AS DGB
            ,DVAMT                                          AS S
        FROM sa01
    """)

    # ── SA12: sa02 변환 ───────────────────────
    _exec(con, "SA12 (sa02 변환)", """
        CREATE OR REPLACE TABLE sa12 AS
        SELECT
             SLPDT                                          AS YYYYMM
            ,NTACC_CD
            ,BZCS_TPCD
            ,RR_ORGCD
            ,SL_TPCD
            ,INS_IMCD
            ,CASE WHEN CVRCD = '*' THEN '0' ELSE CVRCD END AS DGB
            ,DVAMT                                          AS S
        FROM sa02
    """)

    # ── SA20: SA11 + SA12 UNION ALL ───────────
    _exec(con, "SA20 (SA11 + SA12)", """
        CREATE OR REPLACE TABLE sa20 AS
        SELECT * FROM sa11
        UNION ALL
        SELECT * FROM sa12
    """)

    # ── SA_집계: 최종 배분 집계 ───────────────
    _exec(con, f"SA_{yyyymm} (최종 배분집계)", f"""
        CREATE OR REPLACE TABLE sa_{yyyymm} AS
        SELECT DISTINCT
             YYYYMM
            ,NTACC_CD
            ,RR_ORGCD
            ,BZCS_TPCD
            ,SL_TPCD
            ,INS_IMCD
            ,DGB        AS CVRCD
            ,SUM(S)     AS S
        FROM sa20
        GROUP BY YYYYMM, NTACC_CD, RR_ORGCD, BZCS_TPCD, SL_TPCD, INS_IMCD, DGB
    """)

    # ══════════════════════════════════════════════
    # 이연 계산 (EY.A2601 원천)
    # ⚠ ey_a2601 테이블이 TABLE_REGISTRY에 등록되어 있어야 함
    # ══════════════════════════════════════════════

    # ── 이연액: 신계약비이연액 (RL_CREW_NWCRT * -1) ──
    _exec(con, "tmp_TEMP01_01 (이연액 계정매핑)", """
        CREATE OR REPLACE TABLE tmp_TEMP01_01 AS
        SELECT
            '일반관리비'                                        AS GB
           ,CASE
                WHEN SL_TPCD IN ('11','21','31','42') THEN '5125030001'
                WHEN SL_TPCD = '41'                   THEN '5125010001'
                WHEN SL_TPCD IN ('12','13','22','32') THEN '5125050001'
            END                                                AS NTACC_CD
           ,CLS_YYMM                                           AS YYYYMM
           ,'LA'                                               AS BGB
           ,'0'                                                AS CGB
           ,RL_CREW_NWCRT * -1                                 AS S
        FROM ey_a2601
        WHERE SL_TPCD IN ('11','21','31','41','42','12','13','22','32')
    """)

    # ── 상각비: 신계약비상각비 (RL_NWCRT_DPCS) ───────
    _exec(con, "tmp_TEMP01_02 (상각비 계정매핑)", """
        CREATE OR REPLACE TABLE tmp_TEMP01_02 AS
        SELECT
            '일반관리비'                                        AS GB
           ,CASE
                WHEN SL_TPCD IN ('11','21','31','42') THEN '5131030001'
                WHEN SL_TPCD = '41'                   THEN '5131010001'
                WHEN SL_TPCD IN ('12','13','22','32') THEN '5131050001'
            END                                                AS NTACC_CD
           ,CLS_YYMM                                           AS YYYYMM
           ,'LA'                                               AS BGB
           ,'0'                                                AS CGB
           ,RL_NWCRT_DPCS                                      AS S
        FROM ey_a2601
        WHERE SL_TPCD IN ('11','21','31','41','42','12','13','22','32')
    """)

    # ── 이연액 + 상각비 합산 집계 ────────────────────
    _exec(con, "TEMP01 (이연 최종 집계)", """
        CREATE OR REPLACE TABLE TEMP01 AS
        SELECT DISTINCT YYYYMM, GB, NTACC_CD, BGB, CGB, SUM(S) AS S
        FROM (
            SELECT * FROM tmp_TEMP01_01
            UNION ALL
            SELECT * FROM tmp_TEMP01_02
        )
        GROUP BY YYYYMM, GB, NTACC_CD, BGB, CGB
    """)

    # tmp_ 중간 테이블 정리
    con.execute("DROP TABLE IF EXISTS tmp_TEMP01_01")
    con.execute("DROP TABLE IF EXISTS tmp_TEMP01_02")

    # ══════════════════════════════════════════════
    # 순사업비 요약 (출력 대상 → out_ 접두사)
    # ══════════════════════════════════════════════

    # ── out_SA000: 보종별 순사업비 요약 ──────────
    # ⚠ S3GUB, B_CH 컬럼은 KEY.ACCGB 조인 결과에서 옴
    #   key_accgb 테이블이 TABLE_REGISTRY에 등록되어 있어야 함
    _exec(con, "out_SA000 (순사업비 요약)", f"""
        CREATE OR REPLACE TABLE out_SA000 AS
        SELECT DISTINCT
             INS_IMCD
            ,CASE
                WHEN SUBSTR(INS_IMCD,1,2) = 'CA' THEN '자동차'
                WHEN SUBSTR(INS_IMCD,1,2) IN ('FA','MA') THEN '일반'
                WHEN INS_IMCD IN (
                    'LA02660','LA02661','LA02662','LA02663','LA02664','LA02665',
                    'LA02666','LA02667','LA02668','LA02669','LA02670','LA02699',
                    'LA02800','LA02901','LA02802','LA02803','LA02804','LA02805',
                    'LA02806','LA02807','LA02808','LA02809','LA02810','LA02812',
                    'LA02813','LA02815','LA02814','LA02816','LA02818','LA02819',
                    'LA02820','LA02821','LA02822','LA02823','LA02824','LA02827',
                    'LA02825','LA02826','LA02828','LA02829','LA02830','LA02831',
                    'LA02833','LA02835','LA02836','LA02838','LA02839','LA02840',
                    'LA02841','LA02842','LA02844','LA02845','LA02846','LA02847',
                    'LA02848','LA02850','LA02851','LA02852','LA02854','LA02855',
                    'LA02856','LA02858','LA02859','LA02860','LA02861','LA02862',
                    'LA02863','LA02864','LA02867','LA02868','LA02869','LA02852',
                    'LA02866','LA02872','LA02873','LA02874','LA02875','LA02878',
                    'LA02879','LA02880','LA02881','LA02882','LA02883','LA02886',
                    'LA02887','LA02885','LA02888','LA02891','LA02889','LA02890',
                    'LA02892','LA02900','LA02901','LA02910','LA02920','LA02930',
                    'LA02903','LA02904','LA02911','LA02921','LA02931','LA02905',
                    'LA02906','LA02941','LA02951','LA02961','LA02907','LA02908',
                    'LA02971','LA02981','LA02991'
                ) THEN '장기연금'
                WHEN INS_IMCD IN (
                    'LA02660','LA02661','LA02662','LA02663','LA02664','LA02665',
                    'LA02667','LA02668','LA02669','LA02670','LA02699','LA02800',
                    'LA02801','LA02802','LA02803','LA02804','LA02805','LA02806',
                    'LA02807','LA02808','LA02809','LA02810','LA02812','LA02813',
                    'LA02815','LA02814','LA02816','LA02818','LA02819','LA02821',
                    'LA02823','LA02824','LA02827','LA02825','LA02826','LA02828',
                    'LA02829','LA02830','LA02831','LA02833','LA02835','LA02836',
                    'LA02838','LA02839','LA02840','LA02841','LA02842','LA02844',
                    'LA02845','LA02846','LA02847','LA02848','LA02850','LA02851',
                    'LA02852','LA02854','LA02855','LA02856','LA02858','LA02859',
                    'LA02860','LA02861','LA02862','LA02863','LA02864','LA02867',
                    'LA02868','LA02852','LA02866','LA02872','LA02873','LA02874',
                    'LA02875','LA02878','LA02879','LA02880','LA02881','LA02882',
                    'LA02883','LA02886','LA02885','LA02888','LA02891','LA02889',
                    'LA02890','LA02892','LA02900','LA02901','LA02910','LA02920',
                    'LA02930','LA02903','LA02911','LA02921','LA02931','LA02905',
                    'LA02906','LA02941','LA02951','LA02961','LA02907','LA02908',
                    'LA02971','LA02981','LA02991'
                ) THEN '장기우매당'
                WHEN SUBSTR(INS_IMCD,1,2) = 'RA' THEN '퇴직'
                ELSE '에러'
             END                                               AS BGB
            ,S3GUB
            ,B_CH
            ,SL_TPCD
            ,CASE
                WHEN INS_IMCD = 'CA91001'             THEN '1'
                WHEN SUBSTR(INS_IMCD,1,2) = 'CA'      THEN CVRCD
                ELSE '0'
             END                                               AS CGB
            ,SUM(S)                                            AS S
        FROM sa_{yyyymm}
        WHERE YYYYMM = '{yyyymm}'
          AND BZCS_TPCD <> '02'
        GROUP BY INS_IMCD, BGB, S3GUB, B_CH, SL_TPCD, CGB
    """)

    # ── out_SA001: 신계약비+수금비 추세분석 ──────
    _exec(con, "out_SA001 (신계약비·수금비 추세)", f"""
        CREATE OR REPLACE TABLE out_SA001 AS
        SELECT DISTINCT
             YYYYMM
            ,NTACC_CD
            ,SUBSTR(INS_IMCD,1,2)  AS BGB
            ,SL_TPCD
            ,SUM(S)                AS S
        FROM sa_{yyyymm}
        WHERE SUBSTR(NTACC_CD,1,4) IN ('5121','5127','5128','5129')
        GROUP BY YYYYMM, NTACC_CD, BGB, SL_TPCD
    """)

    log.info(f"[LOGIC] 완료  소요: {time.time()-t0:.1f}초")


# ══════════════════════════════════════════════
# 5. VALIDATE — 검증
# ══════════════════════════════════════════════

def step_validate(con, yyyymm: str):
    """건수 검증 + 시산표 검증"""
    t0 = time.time()
    log.info("=" * 55)
    log.info("[VALIDATE] 검증 시작")
    log.info("=" * 55)

    # ── 결과 테이블 건수 출력 ────
    log.info("── 결과 테이블 건수 ──")
    tables = con.execute(
        "SELECT table_name FROM information_schema.tables "
        "WHERE table_schema = 'main' ORDER BY table_name"
    ).fetchall()

    for (tbl,) in tables:
        if tbl.startswith(PREFIX_VALIDATE):
            continue
        cnt = con.execute(f"SELECT COUNT(*) FROM {tbl}").fetchone()[0]
        log.info(f"  {tbl:25s}: {cnt:>10,}건")

    # ── 시산표 검증 ─────────────────
    log.info("── 시산표 검증 ──")

    _exec(con, "val_TEMP00 (전표 집계)", """
        CREATE OR REPLACE TABLE val_TEMP00 AS
        SELECT DISTINCT '전표'  AS GB
              ,NTACC_CD
              ,SUM(WONCR_POAMT) AS S
        FROM erp
        GROUP BY GB, NTACC_CD
    """)

    _exec(con, "val_TEMP01 (배분결과 집계)", f"""
        CREATE OR REPLACE TABLE val_TEMP01 AS
        SELECT DISTINCT '결과'  AS GB
              ,NTACC_CD
              ,SUM(S)           AS S
        FROM sa_{yyyymm}
        GROUP BY GB, NTACC_CD
    """)

    _exec(con, "val_TEMP02 (전표+결과 비교)", """
        CREATE OR REPLACE TABLE val_TEMP02 AS
        SELECT * FROM val_TEMP00
        UNION ALL
        SELECT * FROM val_TEMP01
    """)

    # ── bs104 IBNR 담보코드 누락 (0건이어야 정상) ──
    _exec(con, "val_PY_IBAMT_CH (bs104 IBNR 담보코드 누락)", """
        CREATE OR REPLACE TABLE val_PY_IBAMT_CH AS
        SELECT *
        FROM bs104
        WHERE IKD_GRPCD = 'LA'
          AND CVRCD IS NULL
    """)

    # ── bs105 IBNR 담보코드 누락 (0건이어야 정상) ──
    _exec(con, "val_PY_RFAMT_CH (bs105 IBNR 담보코드 누락)", """
        CREATE OR REPLACE TABLE val_PY_RFAMT_CH AS
        SELECT *
        FROM bs105
        WHERE IKD_GRPCD = 'LA'
          AND CLM_CVRCD IS NULL
    """)

    # ── 자동차 TM/CM 비비례 배분 누락 체크 (0건이어야 정상) ──
    _exec(con, "val_CA_TMCM_Chk (자동차 TM/CM 비비례 오류)", f"""
        CREATE OR REPLACE TABLE val_CA_TMCM_Chk AS
        SELECT *
        FROM sa_{yyyymm}
        WHERE SUBSTR(BZCS_02_DVCD, 1, 1) = 'E'
          AND NTACC_CD IN (
              SELECT DISTINCT NTACC_CD
              FROM erp
              WHERE SUBSTR(BZCS_02_DVCD, 1, 1) = 'E'
          )
          AND SUBSTR(INS_IMCD, 1, 2) = 'CA'
          AND SL_TPCD = '21'
    """)

    # ── 재마감 자동차 금액 확인 ──────────────────
    _exec(con, "val_CA_2601_Chk (재마감 자동차 금액)", f"""
        CREATE OR REPLACE TABLE val_CA_2601_Chk AS
        SELECT SUM(S) AS TOTAL_S
        FROM sa_{yyyymm}
        WHERE SUBSTR(BZCS_02_DVCD, 1, 1) = 'E'
          AND BZCS_02_DVCD NOT IN ('E07','E08','E09','E10','E11','E12','E13')
          AND NTACC_CD IN (
              SELECT DISTINCT NTACC_CD
              FROM erp
              WHERE SUBSTR(BZCS_02_DVCD, 1, 1) = 'E'
          )
          AND SUBSTR(INS_IMCD, 1, 2) = 'CA'
          AND SL_TPCD = '14'
    """)
    total_s = con.execute("SELECT TOTAL_S FROM val_CA_2601_Chk").fetchone()[0]
    log.info(f"  val_CA_2601_Chk TOTAL_S = {total_s:,.0f}  "
             f"(기준: 재작업전 200,449,694 / 재작업후 107,627,489)")

    # ── val_ 임시 테이블 정리 ───────────
    val_tables = [r[0] for r in con.execute(
        "SELECT table_name FROM information_schema.tables "
        "WHERE table_schema='main'"
    ).fetchall() if r[0].startswith(PREFIX_VALIDATE)]
    for t in val_tables:
        con.execute(f"DROP TABLE IF EXISTS {t}")
    if val_tables:
        log.info(f"임시 테이블 정리: {val_tables}")

    log.info(f"[VALIDATE] 완료  소요: {time.time()-t0:.1f}초")


# ══════════════════════════════════════════════
# 6. EXPORT — Excel 출력
# ══════════════════════════════════════════════

def step_export(con, yyyymm: str, tables: list = None):
    """결과 테이블을 Excel로 출력"""
    t0 = time.time()
    out_dir = ROOT / "output"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / f"output_{yyyymm}.xlsx"

    log.info("=" * 55)
    log.info(f"[EXPORT] Excel 저장 시작  → {out_file}")
    log.info("=" * 55)

    # ── 고정 시트맵 ─────────────────────────
    sheet_map = {
        "fio841"          : "fio841",
        "fio843"          : "fio843",
        "erp"             : "ERP_BS100",
        "sa01"            : "SA01_유지비배분",
        "sa02"            : "SA02_손해조사비배분",
        "BASE_DATA_CH"    : "집계_841보험료",
        "BASE_DATA_CH_843": "집계_843보험료",
        "BASE_DATA_CH2"   : "집계_843건수",
        "CH"              : "취급기관이상체크",
        "bs101"           : "BS101_원수실적",
        "bs104"           : "BS104_지급보험금",
        "bs105"           : "BS105_지급준비금",
    }

    # ── out_ 테이블 자동 수집 ──────────
    db_tables = [r[0] for r in con.execute(
        "SELECT table_name FROM information_schema.tables "
        "WHERE table_schema='main' ORDER BY table_name"
    ).fetchall()]
    for tbl in db_tables:
        if tbl.startswith(PREFIX_EXPORT):
            sheet_map[tbl] = tbl[len(PREFIX_EXPORT):]

    if tables:
        sheet_map = {k: v for k, v in sheet_map.items() if k in tables}

    # ── 시트 저장 ──────────────────────
    summary_rows = []

    with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
        for tbl, sheet in sheet_map.items():
            try:
                ts = time.time()
                df_out = con.execute(f"SELECT * FROM {tbl}").df()
                sheet_name = sheet[:31]
                df_out.to_excel(writer, sheet_name=sheet_name, index=False)
                elapsed_s = time.time() - ts
                summary_rows.append((tbl, sheet_name, len(df_out), elapsed_s))
                log.info(f"  시트 저장: {sheet:25s}  {len(df_out):>10,}건  ({elapsed_s:.1f}초)")
            except Exception as e:
                log.warning(f"  시트 건너뜀: {sheet}  ({e})")

        # ── 요약 시트 ─────────────────────────────────
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = writer.book
        ws_sum = wb.create_sheet("요약", 0)

        hdr_fill  = PatternFill("solid", fgColor="1F4E79")
        hdr_font  = Font(color="FFFFFF", bold=True, size=11)
        alt_fill  = PatternFill("solid", fgColor="EBF3FB")
        link_font = Font(color="1F4E79", underline="single", bold=False)
        center    = Alignment(horizontal="center", vertical="center")
        left      = Alignment(horizontal="left",   vertical="center")
        right     = Alignment(horizontal="right",  vertical="center")
        thin      = Side(style="thin", color="BFBFBF")
        border    = Border(left=thin, right=thin, top=thin, bottom=thin)

        # 타이틀
        ws_sum.merge_cells("A1:E1")
        title_cell = ws_sum["A1"]
        title_cell.value = (f"사업비배분 파이프라인 출력 요약  |  {yyyymm}  |  "
                            f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        title_cell.font      = Font(bold=True, size=13, color="1F4E79")
        title_cell.alignment = left
        ws_sum.row_dimensions[1].height = 28

        # 헤더
        headers = ["No", "테이블명", "시트명", "건수", "소요(초)"]
        col_widths = [6, 28, 28, 16, 10]
        for col, (h, w) in enumerate(zip(headers, col_widths), 1):
            c = ws_sum.cell(row=2, column=col, value=h)
            c.fill      = hdr_fill
            c.font      = hdr_font
            c.alignment = center
            c.border    = border
            ws_sum.column_dimensions[get_column_letter(col)].width = w
        ws_sum.row_dimensions[2].height = 20

        # 데이터 행
        for i, (tbl, sheet_name, cnt, el) in enumerate(summary_rows, 1):
            row = i + 2
            fill = alt_fill if i % 2 == 0 else PatternFill()

            c_no = ws_sum.cell(row=row, column=1, value=i)
            c_no.alignment, c_no.border, c_no.fill = center, border, fill

            c_tbl = ws_sum.cell(row=row, column=2, value=tbl)
            c_tbl.alignment, c_tbl.border, c_tbl.fill = left, border, fill

            c_sheet = ws_sum.cell(row=row, column=3, value=sheet_name)
            c_sheet.hyperlink = f"#{sheet_name}!A1"
            c_sheet.font, c_sheet.alignment, c_sheet.border, c_sheet.fill = link_font, left, border, fill

            c_cnt = ws_sum.cell(row=row, column=4, value=cnt)
            c_cnt.number_format, c_cnt.alignment, c_cnt.border, c_cnt.fill = "#,##0", right, border, fill

            c_el = ws_sum.cell(row=row, column=5, value=round(el, 1))
            c_el.alignment, c_el.border, c_el.fill = center, border, fill

            ws_sum.row_dimensions[row].height = 18

        # 합계 행
        sum_row = len(summary_rows) + 3
        total_cnt = sum(r[2] for r in summary_rows)
        ws_sum.merge_cells(f"A{sum_row}:C{sum_row}")
        c_total_label = ws_sum.cell(row=sum_row, column=1, value="합계")
        c_total_label.font, c_total_label.alignment, c_total_label.border = Font(bold=True), center, border
        c_total = ws_sum.cell(row=sum_row, column=4, value=total_cnt)
        c_total.number_format, c_total.font, c_total.alignment, c_total.border = "#,##0", Font(bold=True), right, border
        ws_sum.cell(row=sum_row, column=5).border = border
        ws_sum.row_dimensions[sum_row].height = 20

        log.info(f"  요약 시트 생성: {len(summary_rows)}개 시트  총 {total_cnt:,}건")

    log.info(f"[EXPORT] 완료  → {out_file}  소요: {time.time()-t0:.1f}초")


# ══════════════════════════════════════════════
# 7. CLI
# ══════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="SAS → DuckDB 사업비배분 파이프라인",
        formatter_class=argparse.RawTextHelpFormatter,
        epilog="""
예시:
  python "sas_to_duckdb 12.py" --ym 202601 --stage load logic validate export
  python "sas_to_duckdb 12.py" --ym 202601 --stage load --tables erp sa01
  python "sas_to_duckdb 12.py" --ym 202601 --stage logic validate export
  python "sas_to_duckdb 12.py" --list
        """
    )

    parser.add_argument("--ym", type=str, default="202601",
                        help="처리할 년월 (기본: 202601)")
    parser.add_argument("--stage", "-s", nargs="+",
                        choices=["load", "logic", "validate", "export"],
                        help="실행할 단계 (순서대로 나열)")
    parser.add_argument("--tables", "-t", nargs="+", default=None,
                        help="load/export 시 대상 테이블 지정")
    parser.add_argument("--verbose", "-v", action="store_true",
                        help="DEBUG 로그를 콘솔에도 출력")
    parser.add_argument("--list", "-l", action="store_true",
                        help="등록된 테이블 목록 출력 후 종료")

    args = parser.parse_args()
    yyyymm = args.ym

    if args.verbose:
        for handler in log.handlers:
            if isinstance(handler, logging.StreamHandler) and \
               not isinstance(handler, logging.FileHandler):
                handler.setLevel(logging.DEBUG)

    if args.list:
        log.info(f"등록된 테이블 ({len(TABLE_REGISTRY)}개):")
        for name, cfg in TABLE_REGISTRY.items():
            log.info(f"  {name:10s}  [{cfg['type']:4s}]  {cfg['desc']}")
            log.info(f"             파일: {cfg['file']}")
        return

    if not args.stage:
        log.info("실행할 단계를 지정하세요.  예: --stage load logic validate export")
        return

    # 테이블명 유효성 검사
    if args.tables:
        invalid = [t for t in args.tables if t not in TABLE_REGISTRY]
        if invalid:
            log.error(f"알 수 없는 테이블: {invalid}")
            log.error(f"사용 가능: {list(TABLE_REGISTRY.keys())}")
            return

    # DB 연결
    db_path = ROOT / "db" / f"{yyyymm}.duckdb"
    db_path.parent.mkdir(parents=True, exist_ok=True)
    log.info(f"대상 월: {yyyymm}")
    log.info(f"DB    : {db_path}")

    con = duckdb.connect(str(db_path))
    try:
        t_total = time.time()

        for stage in args.stage:
            if stage == "load":
                step_load(con, yyyymm, tables=args.tables)
            elif stage == "logic":
                step_logic(con, yyyymm)
            elif stage == "validate":
                step_validate(con, yyyymm)
            elif stage == "export":
                step_export(con, yyyymm, tables=args.tables)

        log.info(f"전체 완료  총 소요: {time.time()-t_total:.1f}초")
    finally:
        con.close()


if __name__ == "__main__":
    main()
